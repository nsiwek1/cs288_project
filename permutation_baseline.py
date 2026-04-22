"""
Permutation Baseline for CycleGAN Dialect Detection
====================================================

Train CycleGAN on ALL 3 communities pooled together, then test on each
community via leave-one-chimp-out cross-validation.

Key design:
  - All calls from a given chimpanzee go entirely into train OR test (no leakage)
  - For each fold, one chimp is held out; model trained on everyone else
  - Cycle reconstruction loss is computed for the held-out chimp's calls
  - Permutation test: shuffle community labels (keeping chimp groupings intact)
    and retrain to build a null distribution

This establishes whether the CycleGAN's reconstruction error truly differs
across communities, or whether any apparent differences are just noise.
"""

import json
import sys
import functools
import warnings
from pathlib import Path
from collections import defaultdict

# Force unbuffered print so we can monitor progress
print = functools.partial(print, flush=True)

# Remove project dir from sys.path to prevent code.py from shadowing
# stdlib 'code' module (circular import issue with torch/pdb).
_project_dir = str(Path(__file__).parent)
while _project_dir in sys.path:
    sys.path.remove(_project_dir)

import numpy as np
import torch
import torch.nn as nn
import torch.optim as optim
from sklearn.preprocessing import StandardScaler
from torch.utils.data import DataLoader, TensorDataset
from scipy import stats
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import openpyxl

warnings.filterwarnings("ignore")

# Restore project dir now that torch is imported
sys.path.insert(0, _project_dir)

# ---------------------------------------------------------------------------
# Configuration (mirrored from code.py to avoid import issues)
# ---------------------------------------------------------------------------
BASE_DIR = Path("/Users/natalia_mac/Downloads/Chimpanzee Pant-Hoots")
EMBEDDING_CACHE = BASE_DIR / "embeddings_cache.npz"
RESULTS_DIR = BASE_DIR / "results"
RESULTS_DIR.mkdir(exist_ok=True)

EMBEDDING_DIM = 768
HIDDEN_DIM = 512
N_HIDDEN_LAYERS = 4
LEARNING_RATE = 5e-5
WEIGHT_DECAY = 1e-6
N_EPOCHS = 200
CYCLE_WEIGHT = 20.0
IDENTITY_WEIGHT = 2.0
ADVERSARIAL_WEIGHT = 1.0
BATCH_SIZE = 8
PATIENCE = 30

DEVICE = "mps" if torch.backends.mps.is_available() else "cpu"

COMMUNITY_MAP = {
    "KSK": {"name": "Kasekela", "callers": ["FND", "FO", "FU", "SL", "SN", "ZS"]},
    "MTB": {"name": "Mitumba",  "callers": ["EDG", "FAN", "KOC", "LAM", "LON"]},
    "KAN": {"name": "Kanyawara","callers": ["BB", "ES", "KK", "LK", "PG", "ST", "TJ"]},
}

ALL_CALLERS = {}
for comm, info in COMMUNITY_MAP.items():
    for c in info["callers"]:
        ALL_CALLERS[c] = comm

OUT_DIR = RESULTS_DIR / "permutation_baseline"
OUT_DIR.mkdir(exist_ok=True)

N_PERMUTATIONS = 15


# ---------------------------------------------------------------------------
# CycleGAN Architecture (same as code.py)
# ---------------------------------------------------------------------------
class Generator(nn.Module):
    def __init__(self, input_dim=EMBEDDING_DIM, output_dim=EMBEDDING_DIM,
                 hidden_dim=HIDDEN_DIM, n_layers=N_HIDDEN_LAYERS):
        super().__init__()
        layers = []
        prev_dim = input_dim
        for i in range(n_layers):
            layers.extend([
                nn.Linear(prev_dim, hidden_dim),
                nn.LayerNorm(hidden_dim),
                nn.LeakyReLU(0.2),
                nn.Dropout(0.05),
            ])
            prev_dim = hidden_dim
        layers.append(nn.Linear(prev_dim, output_dim))
        self.net = nn.Sequential(*layers)
        self.residual = (input_dim == output_dim)

    def forward(self, x):
        out = self.net(x)
        if self.residual:
            return x + out
        return out


class Discriminator(nn.Module):
    def __init__(self, input_dim=EMBEDDING_DIM, hidden_dim=HIDDEN_DIM):
        super().__init__()
        self.net = nn.Sequential(
            nn.Linear(input_dim, hidden_dim),
            nn.LeakyReLU(0.2),
            nn.Dropout(0.05),
            nn.Linear(hidden_dim, hidden_dim // 2),
            nn.LeakyReLU(0.2),
            nn.Linear(hidden_dim // 2, 1),
        )

    def forward(self, x):
        return self.net(x)


class CycleGAN:
    def __init__(self, device=DEVICE):
        self.device = device
        self.G_p2h = Generator().to(device)
        self.G_h2p = Generator().to(device)
        self.D_h = Discriminator().to(device)
        self.D_p = Discriminator().to(device)

        self.opt_G = optim.AdamW(
            list(self.G_p2h.parameters()) + list(self.G_h2p.parameters()),
            lr=LEARNING_RATE, weight_decay=WEIGHT_DECAY, betas=(0.5, 0.999))
        self.opt_D = optim.AdamW(
            list(self.D_h.parameters()) + list(self.D_p.parameters()),
            lr=LEARNING_RATE, weight_decay=WEIGHT_DECAY, betas=(0.5, 0.999))

        self.sched_G = optim.lr_scheduler.CosineAnnealingLR(self.opt_G, T_max=N_EPOCHS)
        self.sched_D = optim.lr_scheduler.CosineAnnealingLR(self.opt_D, T_max=N_EPOCHS)

        self.mse = nn.MSELoss()
        self.l1 = nn.L1Loss()

    def train_epoch(self, pant_loader, hoot_loader):
        self.G_p2h.train(); self.G_h2p.train()
        self.D_h.train(); self.D_p.train()

        epoch_losses = defaultdict(float)
        n_batches = 0

        for (pants,), (hoots,) in zip(pant_loader, hoot_loader):
            pants = pants.to(self.device)
            hoots = hoots.to(self.device)
            min_bs = min(pants.size(0), hoots.size(0))
            pants, hoots = pants[:min_bs], hoots[:min_bs]

            self.opt_D.zero_grad()
            fake_hoots = self.G_p2h(pants).detach()
            fake_pants = self.G_h2p(hoots).detach()
            real_label = torch.ones(min_bs, 1, device=self.device) * 0.9
            fake_label = torch.zeros(min_bs, 1, device=self.device) + 0.1
            loss_D = (self.mse(self.D_h(hoots), real_label) +
                      self.mse(self.D_h(fake_hoots), fake_label) +
                      self.mse(self.D_p(pants), real_label) +
                      self.mse(self.D_p(fake_pants), fake_label)) * 0.5
            loss_D.backward()
            self.opt_D.step()

            self.opt_G.zero_grad()
            fake_hoots = self.G_p2h(pants)
            cycle_pants = self.G_h2p(fake_hoots)
            fake_pants = self.G_h2p(hoots)
            cycle_hoots = self.G_p2h(fake_pants)

            loss_adv = (self.mse(self.D_h(fake_hoots), real_label) +
                        self.mse(self.D_p(fake_pants), real_label)) * ADVERSARIAL_WEIGHT
            loss_cycle = (self.l1(cycle_pants, pants) +
                          self.l1(cycle_hoots, hoots)) * CYCLE_WEIGHT
            loss_identity = (self.l1(self.G_h2p(pants), pants) +
                             self.l1(self.G_p2h(hoots), hoots)) * IDENTITY_WEIGHT

            loss_G = loss_adv + loss_cycle + loss_identity
            loss_G.backward()
            torch.nn.utils.clip_grad_norm_(
                list(self.G_p2h.parameters()) + list(self.G_h2p.parameters()),
                max_norm=1.0)
            self.opt_G.step()

            epoch_losses["D"] += loss_D.item()
            epoch_losses["G_cycle"] += loss_cycle.item()
            epoch_losses["G_total"] += loss_G.item()
            n_batches += 1

        self.sched_G.step()
        self.sched_D.step()
        return {k: v / max(n_batches, 1) for k, v in epoch_losses.items()}

    @torch.no_grad()
    def compute_cycle_loss(self, embeddings, vtype):
        self.G_p2h.eval(); self.G_h2p.eval()
        emb = torch.tensor(embeddings, dtype=torch.float32).to(self.device)
        if vtype == "pant":
            reconstructed = self.G_h2p(self.G_p2h(emb))
        else:
            reconstructed = self.G_p2h(self.G_h2p(emb))
        return (emb - reconstructed).abs().mean(dim=1).cpu().numpy()


# ---------------------------------------------------------------------------
# KAN filename mapping
# ---------------------------------------------------------------------------
def load_kan_mapping():
    wb = openpyxl.load_workbook(BASE_DIR / "All_communities.xlsx", read_only=True)
    ws = wb["All_communities"]
    kan_map = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if row[1] == "Kanyawara" and row[2] and row[3]:
            fname = str(row[2]).replace(".0", "")
            try:
                fname = str(int(float(fname)))
            except ValueError:
                pass
            caller = str(row[3])
            kan_map[fname] = caller
    wb.close()
    return kan_map


def get_caller_from_filename(filename, community, kan_map):
    if community == "KAN":
        parts = filename.split("_")
        for length in range(1, len(parts)):
            candidate = "_".join(parts[:length])
            if candidate in kan_map:
                return kan_map[candidate]
        return None
    parts = filename.split("_")
    for part in parts[1:]:
        if part in ALL_CALLERS:
            return part
    return None


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------
def load_data():
    kan_map = load_kan_mapping()
    data = np.load(EMBEDDING_CACHE, allow_pickle=True)
    embeddings = data["embeddings"]
    metadata = json.loads(str(data["metadata"]))

    for m in metadata:
        m["caller"] = get_caller_from_filename(m["filename"], m["community"], kan_map)

    caller_data = {}
    for caller, comm in ALL_CALLERS.items():
        pant_mask = np.array([
            m["caller"] == caller and m["vtype"] == "pant" for m in metadata
        ])
        hoot_mask = np.array([
            m["caller"] == caller and m["vtype"] == "hoot" for m in metadata
        ])
        caller_data[caller] = {
            "community": comm,
            "pant_emb": embeddings[pant_mask],
            "hoot_emb": embeddings[hoot_mask],
        }

    return caller_data


def print_inventory(caller_data):
    print("\n  Caller inventory:")
    for comm in ["KSK", "MTB", "KAN"]:
        callers = COMMUNITY_MAP[comm]["callers"]
        print(f"\n  {COMMUNITY_MAP[comm]['name']} ({comm}):")
        for c in callers:
            d = caller_data[c]
            print(f"    {c:>4s}: {len(d['pant_emb'])} pants, {len(d['hoot_emb'])} hoots")


# ---------------------------------------------------------------------------
# Leave-one-chimp-out training & evaluation
# ---------------------------------------------------------------------------
def train_and_evaluate_fold(train_callers, test_caller, caller_data, verbose=True):
    train_pants_list = [caller_data[c]["pant_emb"] for c in train_callers
                        if len(caller_data[c]["pant_emb"]) > 0]
    train_hoots_list = [caller_data[c]["hoot_emb"] for c in train_callers
                        if len(caller_data[c]["hoot_emb"]) > 0]

    if not train_pants_list or not train_hoots_list:
        return None

    train_pants = np.concatenate(train_pants_list)
    train_hoots = np.concatenate(train_hoots_list)

    test_pants = caller_data[test_caller]["pant_emb"]
    test_hoots = caller_data[test_caller]["hoot_emb"]

    if len(test_pants) == 0 and len(test_hoots) == 0:
        return None

    scaler_p = StandardScaler().fit(train_pants)
    scaler_h = StandardScaler().fit(train_hoots)

    train_p_s = scaler_p.transform(train_pants)
    train_h_s = scaler_h.transform(train_hoots)

    bs = min(BATCH_SIZE, len(train_p_s), len(train_h_s))
    pant_loader = DataLoader(
        TensorDataset(torch.tensor(train_p_s, dtype=torch.float32)),
        batch_size=bs, shuffle=True, drop_last=True)
    hoot_loader = DataLoader(
        TensorDataset(torch.tensor(train_h_s, dtype=torch.float32)),
        batch_size=bs, shuffle=True, drop_last=True)

    model = CycleGAN(device=DEVICE)
    best_loss = float("inf")
    patience_ctr = 0

    for epoch in range(N_EPOCHS):
        losses = model.train_epoch(pant_loader, hoot_loader)
        train_cycle = losses["G_cycle"]
        if train_cycle < best_loss:
            best_loss = train_cycle
            patience_ctr = 0
        else:
            patience_ctr += 1
            if patience_ctr >= PATIENCE:
                if verbose:
                    print(f"      Early stop at epoch {epoch+1}")
                break

        if verbose and (epoch + 1) % 200 == 0:
            print(f"      Epoch {epoch+1:4d} | G_cycle: {losses['G_cycle']:.4f}")

    result = {"caller": test_caller, "community": caller_data[test_caller]["community"]}

    if len(test_pants) > 0:
        test_p_s = scaler_p.transform(test_pants)
        pant_losses = model.compute_cycle_loss(test_p_s, "pant")
        result["pant_losses"] = pant_losses
        result["pant_mean"] = pant_losses.mean()
    else:
        result["pant_losses"] = np.array([])
        result["pant_mean"] = np.nan

    if len(test_hoots) > 0:
        test_h_s = scaler_h.transform(test_hoots)
        hoot_losses = model.compute_cycle_loss(test_h_s, "hoot")
        result["hoot_losses"] = hoot_losses
        result["hoot_mean"] = hoot_losses.mean()
    else:
        result["hoot_losses"] = np.array([])
        result["hoot_mean"] = np.nan

    return result


def run_leave_one_chimp_out(caller_data, caller_communities=None, verbose=True):
    all_callers = [c for c in ALL_CALLERS if (
        len(caller_data[c]["pant_emb"]) > 0 or len(caller_data[c]["hoot_emb"]) > 0
    )]

    if caller_communities is None:
        caller_communities = {c: caller_data[c]["community"] for c in all_callers}

    results = []
    for i, test_caller in enumerate(all_callers):
        train_callers = [c for c in all_callers if c != test_caller]
        if verbose:
            comm = caller_communities[test_caller]
            print(f"\n    Fold {i+1}/{len(all_callers)}: "
                  f"hold out {test_caller} ({comm}), "
                  f"train on {len(train_callers)} chimps")

        fold_result = train_and_evaluate_fold(
            train_callers, test_caller, caller_data, verbose=verbose)

        if fold_result is not None:
            fold_result["community"] = caller_communities[test_caller]
            results.append(fold_result)

    return results


# ---------------------------------------------------------------------------
# Aggregate results by community
# ---------------------------------------------------------------------------
def aggregate_by_community(results):
    community_losses = defaultdict(lambda: {"pant": [], "hoot": []})

    for r in results:
        comm = r["community"]
        if len(r["pant_losses"]) > 0:
            community_losses[comm]["pant"].extend(r["pant_losses"].tolist())
        if len(r["hoot_losses"]) > 0:
            community_losses[comm]["hoot"].extend(r["hoot_losses"].tolist())

    agg = {}
    for comm in ["KSK", "MTB", "KAN"]:
        pant = np.array(community_losses[comm]["pant"])
        hoot = np.array(community_losses[comm]["hoot"])
        combined = np.concatenate([pant, hoot]) if len(pant) + len(hoot) > 0 else np.array([])
        agg[comm] = {
            "pant": pant,
            "hoot": hoot,
            "combined": combined,
            "pant_mean": pant.mean() if len(pant) > 0 else np.nan,
            "hoot_mean": hoot.mean() if len(hoot) > 0 else np.nan,
            "combined_mean": combined.mean() if len(combined) > 0 else np.nan,
        }
    return agg


def compute_observed_stat(agg):
    return agg["KAN"]["combined_mean"] - agg["KSK"]["combined_mean"]


# ---------------------------------------------------------------------------
# Permutation test
# ---------------------------------------------------------------------------
def run_permutation_test(caller_data, observed_stat, n_perms=N_PERMUTATIONS):
    all_callers = [c for c in ALL_CALLERS if (
        len(caller_data[c]["pant_emb"]) > 0 or len(caller_data[c]["hoot_emb"]) > 0
    )]
    real_labels = [caller_data[c]["community"] for c in all_callers]

    rng = np.random.default_rng(42)
    null_stats = []

    for perm_i in range(n_perms):
        shuffled_labels = rng.permutation(real_labels)
        perm_communities = dict(zip(all_callers, shuffled_labels))

        print(f"\n  === Permutation {perm_i+1}/{n_perms} ===")
        perm_results = run_leave_one_chimp_out(
            caller_data, caller_communities=perm_communities, verbose=False)

        perm_agg = aggregate_by_community(perm_results)
        perm_stat = compute_observed_stat(perm_agg)
        null_stats.append(perm_stat)
        print(f"    Stat (KAN-KSK combined): {perm_stat:.4f}")

    null_stats = np.array(null_stats)
    p_value = (np.sum(null_stats >= observed_stat) + 1) / (n_perms + 1)

    return null_stats, p_value


# ---------------------------------------------------------------------------
# Visualization
# ---------------------------------------------------------------------------
def plot_results(agg, per_chimp_results):
    fig, axes = plt.subplots(1, 3, figsize=(16, 5))

    for idx, (vtype, title) in enumerate([
        ("pant", "Pant Cycle Loss"),
        ("hoot", "Hoot Cycle Loss"),
        ("combined", "Combined Cycle Loss"),
    ]):
        ax = axes[idx]
        data_to_plot = []
        labels = []
        colors = ["#2ecc71", "#f39c12", "#e74c3c"]

        for comm, color in zip(["KSK", "MTB", "KAN"], colors):
            losses = agg[comm][vtype]
            data_to_plot.append(losses if len(losses) > 0 else np.array([0]))
            n = len(losses)
            labels.append(f"{comm}\n({COMMUNITY_MAP[comm]['name']})\nn={n}")

        bp = ax.boxplot(data_to_plot, labels=labels, patch_artist=True,
                        widths=0.6, showfliers=True,
                        flierprops=dict(marker="o", markersize=3, alpha=0.5))
        for patch, color in zip(bp["boxes"], colors):
            patch.set_facecolor(color)
            patch.set_alpha(0.6)

        ax.set_title(title, fontsize=11)
        ax.set_ylabel("L1 Reconstruction Error")
        ax.grid(axis="y", alpha=0.3)

        means = [d.mean() for d in data_to_plot]
        ax.scatter(range(1, 4), means, color="black", marker="D", s=40, zorder=3)

    plt.suptitle("Permutation Baseline: Cycle Loss by Community\n"
                 "(Trained on all 3 communities, leave-one-chimp-out CV)",
                 fontsize=13, fontweight="bold")
    plt.tight_layout()
    plt.savefig(OUT_DIR / "permutation_baseline_boxplots.png", dpi=150, bbox_inches="tight")
    plt.close()
    print(f"  Saved: {OUT_DIR / 'permutation_baseline_boxplots.png'}")

    # Per-chimp bar chart
    fig, ax = plt.subplots(figsize=(14, 5))
    callers_sorted = sorted(per_chimp_results, key=lambda r: (r["community"], r["caller"]))
    names = []
    means = []
    bar_colors = []
    comm_colors = {"KSK": "#2ecc71", "MTB": "#f39c12", "KAN": "#e74c3c"}

    for r in callers_sorted:
        combined = np.concatenate([r["pant_losses"], r["hoot_losses"]])
        if len(combined) == 0:
            continue
        names.append(f"{r['caller']}\n({r['community']})")
        means.append(combined.mean())
        bar_colors.append(comm_colors[r["community"]])

    ax.bar(range(len(names)), means, color=bar_colors, alpha=0.7, edgecolor="black", linewidth=0.5)
    ax.set_xticks(range(len(names)))
    ax.set_xticklabels(names, fontsize=9)
    ax.set_ylabel("Mean Combined Cycle Loss")
    ax.set_title("Per-Chimp Cycle Reconstruction Error\n"
                 "(Leave-one-chimp-out, trained on all communities)",
                 fontweight="bold")
    ax.grid(axis="y", alpha=0.3)

    from matplotlib.patches import Patch
    legend_elements = [Patch(facecolor=c, alpha=0.7, label=f"{k} ({COMMUNITY_MAP[k]['name']})")
                       for k, c in comm_colors.items()]
    ax.legend(handles=legend_elements, loc="upper left")

    plt.tight_layout()
    plt.savefig(OUT_DIR / "per_chimp_cycle_loss.png", dpi=150, bbox_inches="tight")
    plt.close()
    print(f"  Saved: {OUT_DIR / 'per_chimp_cycle_loss.png'}")


def plot_null_distribution(null_stats, observed_stat, p_value):
    fig, ax = plt.subplots(figsize=(8, 5))
    ax.hist(null_stats, bins=30, color="#95a5a6", alpha=0.7, edgecolor="black",
            linewidth=0.5, label="Null distribution")
    ax.axvline(observed_stat, color="#e74c3c", linewidth=2, linestyle="--",
               label=f"Observed = {observed_stat:.4f}")
    ax.set_xlabel("Test Statistic (KAN mean - KSK mean, combined cycle loss)")
    ax.set_ylabel("Count")
    ax.set_title(f"Permutation Test: Community Label Shuffle\n"
                 f"p = {p_value:.4f} ({N_PERMUTATIONS} permutations)",
                 fontweight="bold")
    ax.legend()
    ax.grid(axis="y", alpha=0.3)
    plt.tight_layout()
    plt.savefig(OUT_DIR / "permutation_null_distribution.png", dpi=150, bbox_inches="tight")
    plt.close()
    print(f"  Saved: {OUT_DIR / 'permutation_null_distribution.png'}")


# ---------------------------------------------------------------------------
# Report
# ---------------------------------------------------------------------------
def print_report(agg, per_chimp_results, null_stats, observed_stat, p_value):
    lines = []
    def p(s=""):
        print(s)
        lines.append(s)

    p("=" * 70)
    p("PERMUTATION BASELINE: CYCLEGAN TRAINED ON ALL COMMUNITIES")
    p("=" * 70)
    p()
    p("Method: Train CycleGAN on pooled data from all 3 communities,")
    p("        leave-one-chimp-out cross-validation.")
    p("        All calls from a given chimp are in the same train/test split.")
    p()

    p("--- Cycle Reconstruction Error by Community (Mean +/- SD) ---")
    p()
    p(f"{'Community':<20} {'Pant':<25} {'Hoot':<25} {'Combined':<20}")
    p("-" * 70)
    for comm in ["KSK", "MTB", "KAN"]:
        a = agg[comm]
        pant_str = (f"{a['pant'].mean():.4f} +/- {a['pant'].std():.4f}"
                    if len(a['pant']) > 0 else "N/A")
        hoot_str = (f"{a['hoot'].mean():.4f} +/- {a['hoot'].std():.4f}"
                    if len(a['hoot']) > 0 else "N/A")
        comb_str = (f"{a['combined'].mean():.4f} +/- {a['combined'].std():.4f}"
                    if len(a['combined']) > 0 else "N/A")
        p(f"{comm} ({COMMUNITY_MAP[comm]['name']:<12}) {pant_str:<25} {hoot_str:<25} {comb_str}")

    p()
    p("--- Per-Chimp Results ---")
    p()
    for r in sorted(per_chimp_results, key=lambda x: (x["community"], x["caller"])):
        combined = np.concatenate([r["pant_losses"], r["hoot_losses"]])
        if len(combined) == 0:
            p(f"  {r['caller']:>4s} ({r['community']}): no data")
        else:
            p(f"  {r['caller']:>4s} ({r['community']}): "
              f"combined={combined.mean():.4f} +/- {combined.std():.4f} "
              f"(n_pant={len(r['pant_losses'])}, n_hoot={len(r['hoot_losses'])})")

    p()
    p("--- Pairwise Mann-Whitney U Tests ---")
    p()
    for comm_a, comm_b in [("KSK", "MTB"), ("KSK", "KAN"), ("MTB", "KAN")]:
        a = agg[comm_a]["combined"]
        b = agg[comm_b]["combined"]
        if len(a) > 0 and len(b) > 0:
            u, pval = stats.mannwhitneyu(a, b, alternative="two-sided")
            n1, n2 = len(a), len(b)
            effect = 1 - (2 * u) / (n1 * n2)
            p(f"  {comm_a} vs {comm_b}: U={u:.1f}, p={pval:.4f}, "
              f"effect size={effect:.3f}, "
              f"mean diff={b.mean() - a.mean():.4f}")
        else:
            p(f"  {comm_a} vs {comm_b}: insufficient data")

    p()
    p("--- Permutation Test ---")
    p(f"  Test statistic: KAN mean - KSK mean (combined cycle loss)")
    p(f"  Observed: {observed_stat:.4f}")
    p(f"  Null distribution: mean={null_stats.mean():.4f}, "
      f"std={null_stats.std():.4f}")
    p(f"  p-value: {p_value:.4f} ({N_PERMUTATIONS} permutations)")
    p(f"  Interpretation: {'SIGNIFICANT' if p_value < 0.05 else 'NOT SIGNIFICANT'} "
      f"at alpha=0.05")
    p()
    p("=" * 70)

    report_path = OUT_DIR / "permutation_baseline_report.txt"
    report_path.write_text("\n".join(lines))
    print(f"\nReport saved to {report_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    print("=" * 70)
    print("PERMUTATION BASELINE: CycleGAN on All Communities")
    print("Leave-One-Chimp-Out Cross-Validation")
    print("=" * 70)

    print("\n[1] Loading data...")
    caller_data = load_data()
    print_inventory(caller_data)

    print("\n[2] Running leave-one-chimp-out CV (real labels)...")
    real_results = run_leave_one_chimp_out(caller_data)
    real_agg = aggregate_by_community(real_results)

    print("\n\n  --- Real Label Summary ---")
    for comm in ["KSK", "MTB", "KAN"]:
        a = real_agg[comm]
        if len(a["combined"]) > 0:
            print(f"  {comm}: combined mean={a['combined_mean']:.4f} (n={len(a['combined'])})")

    observed_stat = compute_observed_stat(real_agg)
    print(f"\n  Observed test stat (KAN - KSK combined): {observed_stat:.4f}")

    print("\n[3] Generating plots for real results...")
    plot_results(real_agg, real_results)

    print(f"\n[4] Running permutation test ({N_PERMUTATIONS} permutations)...")
    print("    (This will take a while — each permutation retrains the CycleGAN)")
    null_stats, p_value = run_permutation_test(caller_data, observed_stat)

    plot_null_distribution(null_stats, observed_stat, p_value)

    print("\n[5] Report")
    print_report(real_agg, real_results, null_stats, observed_stat, p_value)

    print("\nDone!")


if __name__ == "__main__":
    main()
