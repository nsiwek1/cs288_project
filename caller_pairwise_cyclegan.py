"""
Per-Caller Pairwise CycleGAN Dialect Analysis with Nearest-Neighbor Audio Retrieval
====================================================================================

For each individual caller:
  1. Train a CycleGAN on that caller's buildups <-> climaxes
  2. Test on other callers from same community (in-group)
  3. Test on callers from each other community (out-group)
  4. Find nearest real audio to each synthetic embedding (NN retrieval)

Callers used:
  KSK (Kasekela):  FND, FO, FU, SL, SN, ZS
  MTB (Mitumba):   EDG, FAN, KOC, LAM, LON
  KAN (Kanyawara): BB, ES, KK, LK, PG, ST, TJ
"""

import json
import sys
import importlib.util
import warnings
from pathlib import Path
from collections import defaultdict

# Prevent code.py from shadowing stdlib 'code'
_project_dir = str(Path(__file__).parent)
if _project_dir in sys.path:
    sys.path.remove(_project_dir)

import numpy as np
import torch
from sklearn.preprocessing import StandardScaler
from torch.utils.data import DataLoader, TensorDataset
from scipy import stats
from scipy.spatial.distance import cdist
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import openpyxl

sys.path.insert(0, _project_dir)
_spec = importlib.util.spec_from_file_location(
    "chimp_code", str(Path(__file__).parent / "code.py"))
_mod = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_mod)

EMBEDDING_CACHE = _mod.EMBEDDING_CACHE
RESULTS_DIR = _mod.RESULTS_DIR
DEVICE = _mod.DEVICE
N_EPOCHS = _mod.N_EPOCHS
EMBEDDING_DIM = _mod.EMBEDDING_DIM
HIDDEN_DIM = _mod.HIDDEN_DIM
N_HIDDEN_LAYERS = _mod.N_HIDDEN_LAYERS
LEARNING_RATE = _mod.LEARNING_RATE
WEIGHT_DECAY = _mod.WEIGHT_DECAY
CYCLE_WEIGHT = _mod.CYCLE_WEIGHT
IDENTITY_WEIGHT = _mod.IDENTITY_WEIGHT
ADVERSARIAL_WEIGHT = _mod.ADVERSARIAL_WEIGHT
Discriminator = _mod.Discriminator

import torch.nn as nn
import torch.optim as optim


class GeneratorNoResidual(nn.Module):
    """MLP generator WITHOUT residual connection — forces non-trivial mapping."""
    def __init__(self, input_dim=EMBEDDING_DIM, output_dim=EMBEDDING_DIM,
                 hidden_dim=HIDDEN_DIM, n_layers=N_HIDDEN_LAYERS):
        super().__init__()
        layers = []
        prev_dim = input_dim
        for i in range(n_layers):
            layers.extend([
                nn.Linear(prev_dim, hidden_dim),
                nn.LayerNorm(hidden_dim),
                nn.LeakyReLU(0.2),
                nn.Dropout(0.1),
            ])
            prev_dim = hidden_dim
        layers.append(nn.Linear(prev_dim, output_dim))
        self.net = nn.Sequential(*layers)

    def forward(self, x):
        return self.net(x)  # NO residual — must learn the full mapping


class CycleGAN:
    """Cycle-consistent GAN using generators WITHOUT residual connections."""

    def __init__(self, device=DEVICE):
        self.device = device
        self.G_p2h = GeneratorNoResidual().to(device)
        self.G_h2p = GeneratorNoResidual().to(device)
        self.D_h = Discriminator().to(device)
        self.D_p = Discriminator().to(device)

        self.opt_G = optim.AdamW(
            list(self.G_p2h.parameters()) + list(self.G_h2p.parameters()),
            lr=LEARNING_RATE, weight_decay=WEIGHT_DECAY, betas=(0.5, 0.999))
        self.opt_D = optim.AdamW(
            list(self.D_h.parameters()) + list(self.D_p.parameters()),
            lr=LEARNING_RATE, weight_decay=WEIGHT_DECAY, betas=(0.5, 0.999))

        self.sched_G = optim.lr_scheduler.CosineAnnealingLR(self.opt_G, T_max=N_EPOCHS)
        self.sched_D = optim.lr_scheduler.CosineAnnealingLR(self.opt_D, T_max=N_EPOCHS)

        self.mse = nn.MSELoss()
        self.l1 = nn.L1Loss()

    def train_epoch(self, pant_loader, hoot_loader):
        self.G_p2h.train(); self.G_h2p.train()
        self.D_h.train(); self.D_p.train()

        epoch_losses = defaultdict(float)
        n_batches = 0

        for (pants,), (hoots,) in zip(pant_loader, hoot_loader):
            pants = pants.to(self.device)
            hoots = hoots.to(self.device)
            min_bs = min(pants.size(0), hoots.size(0))
            pants, hoots = pants[:min_bs], hoots[:min_bs]

            # Train Discriminators
            self.opt_D.zero_grad()
            fake_hoots = self.G_p2h(pants).detach()
            fake_pants = self.G_h2p(hoots).detach()
            real_label = torch.ones(min_bs, 1, device=self.device) * 0.9
            fake_label = torch.zeros(min_bs, 1, device=self.device) + 0.1
            loss_D = (self.mse(self.D_h(hoots), real_label) +
                      self.mse(self.D_h(fake_hoots), fake_label) +
                      self.mse(self.D_p(pants), real_label) +
                      self.mse(self.D_p(fake_pants), fake_label)) * 0.5
            loss_D.backward()
            self.opt_D.step()

            # Train Generators
            self.opt_G.zero_grad()
            fake_hoots = self.G_p2h(pants)
            cycle_pants = self.G_h2p(fake_hoots)
            fake_pants = self.G_h2p(hoots)
            cycle_hoots = self.G_p2h(fake_pants)

            loss_adv = (self.mse(self.D_h(fake_hoots), real_label) +
                        self.mse(self.D_p(fake_pants), real_label)) * ADVERSARIAL_WEIGHT
            loss_cycle = (self.l1(cycle_pants, pants) +
                          self.l1(cycle_hoots, hoots)) * CYCLE_WEIGHT
            loss_identity = (self.l1(self.G_h2p(pants), pants) +
                             self.l1(self.G_p2h(hoots), hoots)) * IDENTITY_WEIGHT

            loss_G = loss_adv + loss_cycle + loss_identity
            loss_G.backward()
            torch.nn.utils.clip_grad_norm_(
                list(self.G_p2h.parameters()) + list(self.G_h2p.parameters()),
                max_norm=1.0)
            self.opt_G.step()

            epoch_losses["D"] += loss_D.item()
            epoch_losses["G_adv"] += loss_adv.item()
            epoch_losses["G_cycle"] += loss_cycle.item()
            epoch_losses["G_identity"] += loss_identity.item()
            epoch_losses["G_total"] += loss_G.item()
            n_batches += 1

        self.sched_G.step()
        self.sched_D.step()
        return {k: v / max(n_batches, 1) for k, v in epoch_losses.items()}

    @torch.no_grad()
    def compute_cycle_loss(self, embeddings, vtype):
        self.G_p2h.eval(); self.G_h2p.eval()
        emb = torch.tensor(embeddings, dtype=torch.float32).to(self.device)
        if vtype == "pant":
            reconstructed = self.G_h2p(self.G_p2h(emb))
        else:
            reconstructed = self.G_p2h(self.G_h2p(emb))
        return (emb - reconstructed).abs().mean(dim=1).cpu().numpy()

warnings.filterwarnings("ignore")

BASE_DIR = Path("/Users/natalia_mac/Downloads/Chimpanzee Pant-Hoots")
OUT_DIR = RESULTS_DIR / "caller_pairwise"
OUT_DIR.mkdir(exist_ok=True)

# ---------------------------------------------------------------------------
# Caller definitions
# ---------------------------------------------------------------------------
COMMUNITY_MAP = {
    "KSK": {"name": "Kasekela", "callers": ["FND", "FO", "FU", "SL", "SN", "ZS"]},
    "MTB": {"name": "Mitumba",  "callers": ["EDG", "FAN", "KOC", "LAM", "LON"]},
    "KAN": {"name": "Kanyawara","callers": ["BB", "ES", "KK", "LK", "PG", "ST", "TJ"]},
}

ALL_CALLERS = {}
for comm, info in COMMUNITY_MAP.items():
    for c in info["callers"]:
        ALL_CALLERS[c] = comm


# ---------------------------------------------------------------------------
# KAN filename -> caller mapping from Excel
# ---------------------------------------------------------------------------
def load_kan_mapping():
    """Load mapping from KAN numeric filenames to caller IDs."""
    wb = openpyxl.load_workbook(BASE_DIR / "All_communities.xlsx", read_only=True)
    ws = wb["All_communities"]
    kan_map = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if row[1] == "Kanyawara" and row[2] and row[3]:
            fname = str(row[2]).replace(".0", "")
            try:
                fname = str(int(float(fname)))
            except ValueError:
                pass
            caller = str(row[3])
            kan_map[fname] = caller
    wb.close()
    return kan_map


def get_caller_from_filename(filename, community, kan_map):
    """Extract caller ID from a filename.

    KSK/MTB files: KSK004_FO_BE1 -> FO
    KAN files: 1001_BE4 -> lookup in kan_map -> BB
    """
    if community == "KAN":
        # Strip call type suffix (BE*, CS*) to get the recording ID
        parts = filename.split("_")
        # Try progressively longer prefixes until we find a match
        for length in range(1, len(parts)):
            candidate = "_".join(parts[:length])
            if candidate in kan_map:
                return kan_map[candidate]
        return None

    # KSK/MTB: caller is embedded in filename
    # Patterns: KSK004_FO_BE1, MTB006_FAN_BE2, MTB025_EDG_jan_BE2
    parts = filename.split("_")
    for part in parts[1:]:  # skip recording ID prefix
        if part in ALL_CALLERS:
            return part
    return None


# ---------------------------------------------------------------------------
# Data loading and filtering
# ---------------------------------------------------------------------------
def load_and_organize(kan_map):
    """Load embeddings and organize by caller."""
    data = np.load(EMBEDDING_CACHE, allow_pickle=True)
    embeddings = data["embeddings"]
    metadata = json.loads(str(data["metadata"]))

    # Assign caller to each record
    for m in metadata:
        m["caller"] = get_caller_from_filename(m["filename"], m["community"], kan_map)

    # Build per-caller data
    caller_data = {}
    for caller in ALL_CALLERS:
        comm = ALL_CALLERS[caller]
        pant_mask = np.array([
            m["caller"] == caller and m["vtype"] == "pant"
            for m in metadata
        ])
        hoot_mask = np.array([
            m["caller"] == caller and m["vtype"] == "hoot"
            for m in metadata
        ])
        pant_emb = embeddings[pant_mask]
        hoot_emb = embeddings[hoot_mask]
        pant_meta = [m for m, k in zip(metadata, pant_mask) if k]
        hoot_meta = [m for m, k in zip(metadata, hoot_mask) if k]

        caller_data[caller] = {
            "community": comm,
            "pant_emb": pant_emb,
            "hoot_emb": hoot_emb,
            "pant_meta": pant_meta,
            "hoot_meta": hoot_meta,
        }

    return embeddings, metadata, caller_data


def print_inventory(caller_data):
    """Print data inventory per caller."""
    print("\n  Caller inventory:")
    for comm in ["KSK", "MTB", "KAN"]:
        callers = COMMUNITY_MAP[comm]["callers"]
        print(f"\n  {COMMUNITY_MAP[comm]['name']} ({comm}):")
        for c in callers:
            d = caller_data[c]
            n_p = len(d["pant_emb"])
            n_h = len(d["hoot_emb"])
            flag = " (!) " if min(n_p, n_h) < 3 else "     "
            print(f"    {c:>4s}:{flag}{n_p} pants, {n_h} hoots")


# ---------------------------------------------------------------------------
# Training and evaluation
# ---------------------------------------------------------------------------
def train_caller_model(caller, caller_data):
    """Train a CycleGAN on a single caller's data. Returns model + scalers or None."""
    d = caller_data[caller]
    pants = d["pant_emb"]
    hoots = d["hoot_emb"]

    if len(pants) < 2 or len(hoots) < 2:
        print(f"    {caller}: SKIP (need >=2 pants and >=2 hoots, "
              f"have {len(pants)} pants, {len(hoots)} hoots)")
        return None

    batch_size = min(4, len(pants), len(hoots))

    scaler_p = StandardScaler().fit(pants)
    scaler_h = StandardScaler().fit(hoots)
    p_s = scaler_p.transform(pants)
    h_s = scaler_h.transform(hoots)

    pant_ds = TensorDataset(torch.tensor(p_s, dtype=torch.float32))
    hoot_ds = TensorDataset(torch.tensor(h_s, dtype=torch.float32))
    pant_loader = DataLoader(pant_ds, batch_size=batch_size, shuffle=True, drop_last=True)
    hoot_loader = DataLoader(hoot_ds, batch_size=batch_size, shuffle=True, drop_last=True)

    model = CycleGAN(device=DEVICE)
    for epoch in range(N_EPOCHS):
        losses = model.train_epoch(pant_loader, hoot_loader)

    print(f"    {caller}: trained (batch={batch_size}, "
          f"G_cycle={losses['G_cycle']:.4f})")
    return {"model": model, "scaler_p": scaler_p, "scaler_h": scaler_h}


def evaluate_on_caller(trained_model, target_caller, caller_data):
    """Evaluate a trained model on a target caller's data. Returns cycle losses."""
    model = trained_model["model"]
    scaler_p = trained_model["scaler_p"]
    scaler_h = trained_model["scaler_h"]

    d = caller_data[target_caller]
    pants = d["pant_emb"]
    hoots = d["hoot_emb"]

    result = {"n_pants": len(pants), "n_hoots": len(hoots)}

    if len(pants) > 0:
        pants_s = scaler_p.transform(pants)
        result["pant_losses"] = model.compute_cycle_loss(pants_s, "pant")
    else:
        result["pant_losses"] = np.array([])

    if len(hoots) > 0:
        hoots_s = scaler_h.transform(hoots)
        result["hoot_losses"] = model.compute_cycle_loss(hoots_s, "hoot")
    else:
        result["hoot_losses"] = np.array([])

    all_losses = np.concatenate([
        l for l in [result["pant_losses"], result["hoot_losses"]] if len(l) > 0
    ])
    result["combined_losses"] = all_losses
    result["combined_mean"] = all_losses.mean() if len(all_losses) > 0 else np.nan

    return result


# ---------------------------------------------------------------------------
# Nearest-neighbor audio retrieval
# ---------------------------------------------------------------------------
@torch.no_grad()
def get_synthetic_embeddings(trained_model, caller_data, source_caller):
    """Get the synthetic embeddings produced by cycling a caller's data."""
    model = trained_model["model"]
    scaler_p = trained_model["scaler_p"]
    scaler_h = trained_model["scaler_h"]
    d = caller_data[source_caller]

    synthetics = []

    # pant -> synthetic hoot -> reconstructed pant
    if len(d["pant_emb"]) > 0:
        pants_s = scaler_p.transform(d["pant_emb"])
        emb = torch.tensor(pants_s, dtype=torch.float32).to(model.G_p2h.net[0].weight.device)
        synthetic_hoots = model.G_p2h(emb).cpu().numpy()
        # Un-scale back to original embedding space using hoot scaler
        synthetic_hoots_orig = scaler_h.inverse_transform(synthetic_hoots)
        for i, meta in enumerate(d["pant_meta"]):
            synthetics.append({
                "source_file": meta["path"],
                "source_caller": source_caller,
                "source_vtype": "pant",
                "synthetic_vtype": "hoot",
                "synthetic_emb": synthetic_hoots_orig[i],
            })

    # hoot -> synthetic pant -> reconstructed hoot
    if len(d["hoot_emb"]) > 0:
        hoots_s = scaler_h.transform(d["hoot_emb"])
        emb = torch.tensor(hoots_s, dtype=torch.float32).to(model.G_h2p.net[0].weight.device)
        synthetic_pants = model.G_h2p(emb).cpu().numpy()
        synthetic_pants_orig = scaler_p.inverse_transform(synthetic_pants)
        for i, meta in enumerate(d["hoot_meta"]):
            synthetics.append({
                "source_file": meta["path"],
                "source_caller": source_caller,
                "source_vtype": "hoot",
                "synthetic_vtype": "pant",
                "synthetic_emb": synthetic_pants_orig[i],
            })

    return synthetics


def nearest_neighbor_retrieval(synthetics, all_embeddings, all_metadata, k=3):
    """For each synthetic embedding, find the k nearest real recordings."""
    if not synthetics:
        return []

    synth_embs = np.array([s["synthetic_emb"] for s in synthetics])
    dists = cdist(synth_embs, all_embeddings, metric="cosine")

    results = []
    for i, synth in enumerate(synthetics):
        top_k_idx = np.argsort(dists[i])[:k]
        neighbors = []
        for idx in top_k_idx:
            m = all_metadata[idx]
            neighbors.append({
                "path": m["path"],
                "filename": m["filename"],
                "community": m["community"],
                "caller": m.get("caller", "?"),
                "vtype": m["vtype"],
                "cosine_dist": float(dists[i, idx]),
            })
        results.append({
            "source_file": synth["source_file"],
            "source_caller": synth["source_caller"],
            "source_vtype": synth["source_vtype"],
            "synthetic_vtype": synth["synthetic_vtype"],
            "neighbors": neighbors,
        })
    return results


# ---------------------------------------------------------------------------
# Main analysis
# ---------------------------------------------------------------------------
def run_pairwise_analysis(caller_data, all_embeddings, all_metadata):
    """Train each caller's model and test on all others."""
    all_results = {}       # all_results[train_caller][test_caller] = eval result
    nn_results = {}        # nn_results[train_caller] = list of NN retrievals
    trained_models = {}

    for comm in ["KSK", "MTB", "KAN"]:
        callers = COMMUNITY_MAP[comm]["callers"]
        print(f"\n{'='*60}")
        print(f"  Training models for {COMMUNITY_MAP[comm]['name']} ({comm})")
        print(f"{'='*60}")

        for train_caller in callers:
            print(f"\n  --- {train_caller} ---")
            model_info = train_caller_model(train_caller, caller_data)
            if model_info is None:
                continue

            trained_models[train_caller] = model_info
            all_results[train_caller] = {}

            # Evaluate on ALL callers
            for test_caller in ALL_CALLERS:
                d = caller_data[test_caller]
                if len(d["pant_emb"]) == 0 and len(d["hoot_emb"]) == 0:
                    continue
                result = evaluate_on_caller(model_info, test_caller, caller_data)
                all_results[train_caller][test_caller] = result

            # Nearest-neighbor retrieval for the training caller
            synthetics = get_synthetic_embeddings(model_info, caller_data, train_caller)
            nn = nearest_neighbor_retrieval(synthetics, all_embeddings, all_metadata, k=3)
            nn_results[train_caller] = nn

    return all_results, nn_results, trained_models


# ---------------------------------------------------------------------------
# Aggregation: in-group vs out-group
# ---------------------------------------------------------------------------
def aggregate_results(all_results, caller_data):
    """For each trained model, aggregate in-group vs out-group losses."""
    summary = {}

    for train_caller, test_results in all_results.items():
        train_comm = ALL_CALLERS[train_caller]

        # Collect losses by group
        in_group_losses = []
        out_group_losses = {c: [] for c in ["KSK", "MTB", "KAN"] if c != train_comm}

        for test_caller, result in test_results.items():
            if test_caller == train_caller:
                continue  # skip self
            test_comm = ALL_CALLERS[test_caller]
            losses = result["combined_losses"]
            if len(losses) == 0:
                continue

            if test_comm == train_comm:
                in_group_losses.append(losses)
            else:
                out_group_losses[test_comm].append(losses)

        in_group = np.concatenate(in_group_losses) if in_group_losses else np.array([])
        out_by_comm = {}
        for comm, loss_list in out_group_losses.items():
            out_by_comm[comm] = np.concatenate(loss_list) if loss_list else np.array([])
        all_out = np.concatenate(list(out_by_comm.values())) if any(
            len(v) > 0 for v in out_by_comm.values()) else np.array([])

        summary[train_caller] = {
            "community": train_comm,
            "in_group": in_group,
            "out_group_by_comm": out_by_comm,
            "all_out_group": all_out,
        }

    return summary


# ---------------------------------------------------------------------------
# Statistical tests
# ---------------------------------------------------------------------------
def run_statistics(summary):
    """Run Mann-Whitney U tests: in-group vs out-group for each trained model."""
    stats_out = {}

    for train_caller, s in summary.items():
        in_g = s["in_group"]
        all_out = s["all_out_group"]

        if len(in_g) < 2 or len(all_out) < 2:
            continue

        # In-group vs all out-group
        u, p = stats.mannwhitneyu(in_g, all_out, alternative="less")
        n1, n2 = len(in_g), len(all_out)
        effect = 1 - (2 * u) / (n1 * n2)

        stats_out[train_caller] = {
            "in_vs_all_out": {"U": u, "p": p, "effect": effect,
                              "in_mean": in_g.mean(), "out_mean": all_out.mean()},
        }

        # In-group vs each out community
        for comm, out_losses in s["out_group_by_comm"].items():
            if len(out_losses) < 2:
                continue
            u2, p2 = stats.mannwhitneyu(in_g, out_losses, alternative="less")
            e2 = 1 - (2 * u2) / (len(in_g) * len(out_losses))
            stats_out[train_caller][f"in_vs_{comm}"] = {
                "U": u2, "p": p2, "effect": e2,
                "in_mean": in_g.mean(), "out_mean": out_losses.mean(),
            }

    return stats_out


# ---------------------------------------------------------------------------
# Plotting
# ---------------------------------------------------------------------------
def plot_per_caller_boxplots(all_results, caller_data):
    """For each community, show boxplots of each caller's model tested on all groups."""
    comm_colors = {"KSK": "#2ecc71", "MTB": "#f39c12", "KAN": "#e74c3c"}

    for train_comm in ["KSK", "MTB", "KAN"]:
        callers = [c for c in COMMUNITY_MAP[train_comm]["callers"]
                   if c in all_results]
        if not callers:
            continue

        n_callers = len(callers)
        fig, axes = plt.subplots(1, n_callers, figsize=(5 * n_callers, 6),
                                 squeeze=False)

        for col, train_caller in enumerate(callers):
            ax = axes[0, col]

            # Gather losses by group
            groups = {"self": [], "in-group": [], "KSK": [], "MTB": [], "KAN": []}
            for test_caller, result in all_results[train_caller].items():
                losses = result["combined_losses"]
                if len(losses) == 0:
                    continue
                test_comm = ALL_CALLERS[test_caller]
                if test_caller == train_caller:
                    groups["self"].append(losses)
                elif test_comm == train_comm:
                    groups["in-group"].append(losses)
                else:
                    groups[test_comm].append(losses)

            plot_data = []
            plot_labels = []
            plot_colors = []

            for label, color in [("self", "#3498db"), ("in-group", "#2ecc71")]:
                if groups[label]:
                    combined = np.concatenate(groups[label])
                    plot_data.append(combined)
                    plot_labels.append(f"{label}\nn={len(combined)}")
                    plot_colors.append(color)

            for out_comm in ["KSK", "MTB", "KAN"]:
                if out_comm == train_comm:
                    continue
                if groups[out_comm]:
                    combined = np.concatenate(groups[out_comm])
                    plot_data.append(combined)
                    plot_labels.append(f"{out_comm}\nn={len(combined)}")
                    plot_colors.append(comm_colors[out_comm])

            if not plot_data:
                continue

            bp = ax.boxplot(plot_data, labels=plot_labels, patch_artist=True,
                            widths=0.6,
                            flierprops=dict(marker="o", markersize=3, alpha=0.4))
            for patch, c in zip(bp["boxes"], plot_colors):
                patch.set_facecolor(c)
                patch.set_alpha(0.6)

            # Individual points
            for i, d in enumerate(plot_data):
                x = np.random.normal(i + 1, 0.04, size=len(d))
                ax.scatter(x, d, alpha=0.3, s=8, color=plot_colors[i],
                           edgecolors="black", linewidths=0.2, zorder=3)

            ax.set_title(f"{train_caller} model", fontweight="bold")
            ax.set_ylabel("L1 Cycle Loss")
            ax.grid(axis="y", alpha=0.3)

        plt.suptitle(
            f"{COMMUNITY_MAP[train_comm]['name']} Callers: "
            f"Self vs In-Group vs Out-Group Cycle Loss",
            fontsize=13, fontweight="bold")
        plt.tight_layout()
        path = OUT_DIR / f"boxplots_{train_comm}.png"
        plt.savefig(path, dpi=150, bbox_inches="tight")
        plt.close()
        print(f"  Saved: {path}")


def plot_summary_heatmap(all_results):
    """Heatmap of mean cycle loss: train_caller x test_caller."""
    train_callers = sorted(all_results.keys(),
                           key=lambda c: (["KSK","MTB","KAN"].index(ALL_CALLERS[c]), c))
    test_callers = list(train_callers)  # same set

    matrix = np.full((len(train_callers), len(test_callers)), np.nan)
    for i, tc in enumerate(train_callers):
        for j, tst in enumerate(test_callers):
            if tst in all_results[tc]:
                matrix[i, j] = all_results[tc][tst]["combined_mean"]

    fig, ax = plt.subplots(figsize=(14, 12))
    im = ax.imshow(matrix, cmap="YlOrRd", aspect="equal")

    # Community separators
    comm_boundaries = []
    prev_comm = None
    for i, c in enumerate(train_callers):
        curr = ALL_CALLERS[c]
        if prev_comm and curr != prev_comm:
            comm_boundaries.append(i - 0.5)
        prev_comm = curr

    for b in comm_boundaries:
        ax.axhline(b, color="black", linewidth=2)
        ax.axvline(b, color="black", linewidth=2)

    ax.set_xticks(range(len(test_callers)))
    ax.set_yticks(range(len(train_callers)))
    labels = [f"{c} ({ALL_CALLERS[c]})" for c in train_callers]
    ax.set_xticklabels(labels, rotation=45, ha="right", fontsize=9)
    ax.set_yticklabels(labels, fontsize=9)
    ax.set_xlabel("Test Caller", fontweight="bold")
    ax.set_ylabel("Train Caller", fontweight="bold")
    ax.set_title("Per-Caller CycleGAN: Mean Cycle Loss\n"
                 "(rows=trained on, cols=tested on)",
                 fontweight="bold", fontsize=13)

    # Annotate cells
    for i in range(len(train_callers)):
        for j in range(len(test_callers)):
            if not np.isnan(matrix[i, j]):
                color = "white" if matrix[i, j] > np.nanmean(matrix) else "black"
                ax.text(j, i, f"{matrix[i, j]:.3f}", ha="center", va="center",
                        color=color, fontsize=6)

    plt.colorbar(im, ax=ax, label="Mean L1 Cycle Loss", shrink=0.8)
    plt.tight_layout()
    path = OUT_DIR / "heatmap_caller_pairwise.png"
    plt.savefig(path, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"  Saved: {path}")


def plot_ingroup_vs_outgroup(summary, stats_out):
    """Bar chart: in-group vs out-group mean for each trained caller."""
    callers = sorted(summary.keys(),
                     key=lambda c: (["KSK","MTB","KAN"].index(ALL_CALLERS[c]), c))
    callers = [c for c in callers if len(summary[c]["in_group"]) > 0
               and len(summary[c]["all_out_group"]) > 0]

    fig, ax = plt.subplots(figsize=(max(12, len(callers) * 1.2), 6))
    x = np.arange(len(callers))
    width = 0.35

    in_means = [summary[c]["in_group"].mean() for c in callers]
    out_means = [summary[c]["all_out_group"].mean() for c in callers]
    in_stds = [summary[c]["in_group"].std() for c in callers]
    out_stds = [summary[c]["all_out_group"].std() for c in callers]

    ax.bar(x - width/2, in_means, width, yerr=in_stds, label="In-group",
           color="#2ecc71", alpha=0.7, capsize=3)
    ax.bar(x + width/2, out_means, width, yerr=out_stds, label="Out-group",
           color="#e74c3c", alpha=0.7, capsize=3)

    # Significance stars
    for i, c in enumerate(callers):
        if c in stats_out:
            p = stats_out[c]["in_vs_all_out"]["p"]
            sig = "***" if p < 0.001 else "**" if p < 0.01 else "*" if p < 0.05 else ""
            if sig:
                y = max(in_means[i] + in_stds[i], out_means[i] + out_stds[i]) * 1.05
                ax.text(i, y, sig, ha="center", fontsize=10, fontweight="bold")

    # Community background shading
    comm_colors_bg = {"KSK": "#2ecc7120", "MTB": "#f39c1220", "KAN": "#e74c3c20"}
    prev_comm = None
    start = 0
    for i, c in enumerate(callers + [None]):
        curr = ALL_CALLERS[c] if c else None
        if prev_comm and curr != prev_comm:
            ax.axvspan(start - 0.5, i - 0.5, alpha=0.08,
                       color={"KSK": "#2ecc71", "MTB": "#f39c12", "KAN": "#e74c3c"}[prev_comm])
            mid = (start + i - 1) / 2
            ax.text(mid, ax.get_ylim()[1] * 0.95,
                    COMMUNITY_MAP[prev_comm]["name"],
                    ha="center", fontsize=9, fontstyle="italic", alpha=0.7)
            start = i
        prev_comm = curr

    ax.set_xticks(x)
    ax.set_xticklabels(callers, fontsize=10)
    ax.set_ylabel("Mean L1 Cycle Loss")
    ax.set_title("Per-Caller Model: In-Group vs Out-Group Reconstruction Error\n"
                 "(* p<0.05, ** p<0.01, *** p<0.001, one-sided Mann-Whitney)",
                 fontweight="bold")
    ax.legend()
    ax.grid(axis="y", alpha=0.3)
    plt.tight_layout()
    path = OUT_DIR / "ingroup_vs_outgroup.png"
    plt.savefig(path, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"  Saved: {path}")


# ---------------------------------------------------------------------------
# Report
# ---------------------------------------------------------------------------
def write_report(summary, stats_out, nn_results, all_results):
    """Write text report."""
    lines = []
    def p(s=""):
        lines.append(s)

    p("=" * 70)
    p("PER-CALLER PAIRWISE CYCLEGAN DIALECT ANALYSIS")
    p("=" * 70)
    p()

    # Per-caller summary
    for comm in ["KSK", "MTB", "KAN"]:
        p(f"--- {COMMUNITY_MAP[comm]['name']} ({comm}) ---")
        callers = [c for c in COMMUNITY_MAP[comm]["callers"] if c in summary]
        for c in callers:
            s = summary[c]
            in_mean = s["in_group"].mean() if len(s["in_group"]) > 0 else float("nan")
            out_mean = s["all_out_group"].mean() if len(s["all_out_group"]) > 0 else float("nan")

            p(f"\n  {c} (trained on {len(caller_data_global[c]['pant_emb'])} pants, "
              f"{len(caller_data_global[c]['hoot_emb'])} hoots):")
            p(f"    In-group mean:  {in_mean:.4f} (n={len(s['in_group'])})")
            p(f"    Out-group mean: {out_mean:.4f} (n={len(s['all_out_group'])})")

            for out_comm, out_losses in s["out_group_by_comm"].items():
                if len(out_losses) > 0:
                    p(f"      vs {out_comm}: {out_losses.mean():.4f} (n={len(out_losses)})")

            if c in stats_out:
                st = stats_out[c]["in_vs_all_out"]
                sig = "YES" if st["p"] < 0.05 else "NO"
                p(f"    Mann-Whitney (in < out): U={st['U']:.0f}, p={st['p']:.4f}, "
                  f"effect={st['effect']:.3f} -> significant: {sig}")
        p()

    # Nearest-neighbor examples
    p("=" * 70)
    p("NEAREST-NEIGHBOR AUDIO RETRIEVAL (top examples)")
    p("=" * 70)
    for train_caller in sorted(nn_results.keys()):
        nn_list = nn_results[train_caller]
        if not nn_list:
            continue
        p(f"\n  {train_caller} ({ALL_CALLERS[train_caller]}):")
        for item in nn_list[:6]:  # show first 6
            src = Path(item["source_file"]).name
            p(f"    {src} ({item['source_vtype']}) -> synthetic {item['synthetic_vtype']}:")
            for nb in item["neighbors"]:
                p(f"      -> {nb['filename']} ({nb['community']}/{nb['caller']}, "
                  f"{nb['vtype']}, dist={nb['cosine_dist']:.4f})")
    p()

    # Overall pattern
    p("=" * 70)
    p("OVERALL PATTERN")
    p("=" * 70)
    sig_count = sum(1 for c in stats_out if stats_out[c]["in_vs_all_out"]["p"] < 0.05)
    total = len(stats_out)
    p(f"\n  {sig_count}/{total} callers show significantly lower in-group cycle loss")
    p(f"  (one-sided Mann-Whitney U, p < 0.05)")
    p()

    report_text = "\n".join(lines)
    report_path = OUT_DIR / "caller_pairwise_report.txt"
    report_path.write_text(report_text)
    print(f"\n  Report saved: {report_path}")
    print(report_text)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
caller_data_global = None  # module-level ref for report

def main():
    global caller_data_global

    print("=" * 70)
    print("PER-CALLER PAIRWISE CYCLEGAN ANALYSIS")
    print(f"Device: {DEVICE}")
    print("=" * 70)

    # Load data
    print("\n[1] Loading KAN filename mapping...")
    kan_map = load_kan_mapping()
    print(f"  Mapped {len(kan_map)} KAN filenames to callers")

    print("\n[2] Loading and organizing embeddings...")
    all_embeddings, all_metadata, caller_data = load_and_organize(kan_map)
    caller_data_global = caller_data
    print_inventory(caller_data)

    # Run pairwise analysis
    print("\n[3] Training per-caller CycleGANs and evaluating...")
    all_results, nn_results, trained_models = run_pairwise_analysis(
        caller_data, all_embeddings, all_metadata)

    # Aggregate
    print("\n[4] Aggregating in-group vs out-group...")
    summary = aggregate_results(all_results, caller_data)
    stats_out = run_statistics(summary)

    # Plots
    print("\n[5] Generating plots...")
    plot_per_caller_boxplots(all_results, caller_data)
    plot_summary_heatmap(all_results)
    plot_ingroup_vs_outgroup(summary, stats_out)

    # Report
    print("\n[6] Writing report...")
    write_report(summary, stats_out, nn_results, all_results)

    # Save raw results
    save_data = {}
    for tc in all_results:
        for tst in all_results[tc]:
            key = f"{tc}_on_{tst}"
            losses = all_results[tc][tst]["combined_losses"]
            if len(losses) > 0:
                save_data[key] = losses
    np.savez(OUT_DIR / "raw_caller_pairwise.npz", **save_data)
    print(f"\n  Raw data saved: {OUT_DIR / 'raw_caller_pairwise.npz'}")

    # Save NN results as JSON
    nn_json = {}
    for caller, nn_list in nn_results.items():
        nn_json[caller] = []
        for item in nn_list:
            nn_json[caller].append({
                "source_file": item["source_file"],
                "source_caller": item["source_caller"],
                "source_vtype": item["source_vtype"],
                "synthetic_vtype": item["synthetic_vtype"],
                "neighbors": item["neighbors"],
            })
    with open(OUT_DIR / "nearest_neighbors.json", "w") as f:
        json.dump(nn_json, f, indent=2)
    print(f"  NN results saved: {OUT_DIR / 'nearest_neighbors.json'}")

    print("\nDone!")


if __name__ == "__main__":
    main()
